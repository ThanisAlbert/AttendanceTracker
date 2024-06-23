from datetime import datetime, date

from django.db.models import Q
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.workbook import Workbook

from teams.models import DailyAttendance


class GetReport:

    def __init__(self,fromdate,todate,email):
        self.fromdate = fromdate
        self.todate = todate
        self.email = email


    def Generate(self):

        startdate = datetime.strptime(self.fromdate, "%d/%m/%Y")
        enddate = datetime.strptime(self.todate, "%d/%m/%Y")
        email = self.email

        filtered_records = DailyAttendance.objects.filter(
            Q(attendancedate__gte=startdate) &
            Q(attendancedate__lte=enddate) &
            Q(Tlmail=email)
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "TeamReport"

        ws.merge_cells('A1:I1')
        ws['A1'] = 'ConsolidateReport'
        ws['A1'].font = Font(bold=True)

        fill = PatternFill(start_color="EFEFAF", end_color="EFEFAF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
            for cell in row:
                cell.fill = fill

        fill = PatternFill(start_color="F4F6F6", end_color="F4F6F6", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=9):
            for cell in row:
                cell.fill = fill

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=1, value='Date')
        ws.cell(row=2, column=2, value='Emp ID')
        ws.cell(row=2, column=3, value='EmpName')
        ws.cell(row=2, column=4, value='Team')
        ws.cell(row=2, column=5, value='SAPID')
        ws.cell(row=2, column=6, value='Shift')
        ws.cell(row=2, column=7, value='TLName')
        ws.cell(row=2, column=8, value='Work')
        ws.cell(row=2, column=9, value='Remarks')

        ws['A2'].font = Font(bold=True)
        ws['B2'].font = Font(bold=True)
        ws['C2'].font = Font(bold=True)
        ws['D2'].font = Font(bold=True)
        ws['E2'].font = Font(bold=True)
        ws['F2'].font = Font(bold=True)
        ws['G2'].font = Font(bold=True)
        ws['H2'].font = Font(bold=True)
        ws['I2'].font = Font(bold=True)

        row = 3
        for record in filtered_records:
            ws.cell(row=row, column=1, value=record.attendancedate.strftime("%d/%m/%Y"))
            ws.cell(row=row, column=2, value=record.Employeeno)
            ws.cell(row=row, column=3, value=record.Employeename)
            ws.cell(row=row, column=4, value=record.Teamname)
            ws.cell(row=row, column=5, value=record.Sapid)
            ws.cell(row=row, column=6, value=record.Shift)
            ws.cell(row=row, column=7, value=record.Tlname)
            ws.cell(row=row, column=8, value=record.worktype)
            ws.cell(row=row, column=9, value=record.remarks)
            row = row + 1

        for i in range(1, 8):
            column_number = i
            column = str(chr(64 + column_number))
            ws.column_dimensions[column].width = 18

        column_index = 1  # Column A
        alignment = Alignment(horizontal='left')
        for row in ws.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                cell.alignment = alignment

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        last_row_num = ws.max_row

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        for row in range(2, last_row_num + 1):
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border_style

        today = date.today()
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Report_{}.xlsx"'.format(today.strftime("%d-%m-%Y"))
        wb.save(response)

        return response


    def Generate_satya(self):

        startdate = datetime.strptime(self.fromdate, "%d/%m/%Y")
        enddate = datetime.strptime(self.todate, "%d/%m/%Y")

        filtered_records = DailyAttendance.objects.filter(
            Q(attendancedate__gte=startdate) &
            Q(attendancedate__lte=enddate)
        )

        wb = Workbook()
        ws = wb.active
        ws.title = "TeamReport"

        ws.merge_cells('A1:I1')
        ws['A1'] = 'ConsolidateReport'
        ws['A1'].font = Font(bold=True)

        fill = PatternFill(start_color="EFEFAF", end_color="EFEFAF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
            for cell in row:
                cell.fill = fill

        fill = PatternFill(start_color="F4F6F6", end_color="F4F6F6", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=9):
            for cell in row:
                cell.fill = fill

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=1, value='Date')
        ws.cell(row=2, column=2, value='Emp ID')
        ws.cell(row=2, column=3, value='EmpName')
        ws.cell(row=2, column=4, value='Team')
        ws.cell(row=2, column=5, value='SAPID')
        ws.cell(row=2, column=6, value='Shift')
        ws.cell(row=2, column=7, value='TLName')
        ws.cell(row=2, column=8, value='Work')
        ws.cell(row=2, column=9, value='Remarks')

        ws['A2'].font = Font(bold=True)
        ws['B2'].font = Font(bold=True)
        ws['C2'].font = Font(bold=True)
        ws['D2'].font = Font(bold=True)
        ws['E2'].font = Font(bold=True)
        ws['F2'].font = Font(bold=True)
        ws['G2'].font = Font(bold=True)
        ws['H2'].font = Font(bold=True)
        ws['I2'].font = Font(bold=True)

        row = 3
        for record in filtered_records:
            ws.cell(row=row, column=1, value=record.attendancedate.strftime("%d/%m/%Y"))
            ws.cell(row=row, column=2, value=record.Employeeno)
            ws.cell(row=row, column=3, value=record.Employeename)
            ws.cell(row=row, column=4, value=record.Teamname)
            ws.cell(row=row, column=5, value=record.Sapid)
            ws.cell(row=row, column=6, value=record.Shift)
            ws.cell(row=row, column=7, value=record.Tlname)
            ws.cell(row=row, column=8, value=record.worktype)
            ws.cell(row=row, column=9, value=record.remarks)
            row = row + 1

        for i in range(1, 8):
            column_number = i
            column = str(chr(64 + column_number))
            ws.column_dimensions[column].width = 18

        column_index = 1  # Column A
        alignment = Alignment(horizontal='left')
        for row in ws.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                cell.alignment = alignment

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=6):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        last_row_num = ws.max_row

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        for row in range(2, last_row_num + 1):
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border_style

        today = date.today()
        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Report_{}.xlsx"'.format(today.strftime("%d-%m-%Y"))
        wb.save(response)

        return response




