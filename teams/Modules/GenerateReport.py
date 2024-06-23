from datetime import datetime, date
from django.http import HttpResponse
from openpyxl.styles import PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.workbook import Workbook

from teams.Modules.ShiftTimings import Shift
from teams.models import Login, Team, workon, DailyAttendance, ShiftTimings


class GenerateReport:

    def __init__(self,target_date_str):
        self.target_date_str=target_date_str

    def Generate(self):

        target_date = datetime.strptime(self.target_date_str, "%d/%m/%Y")

        attendance_records = DailyAttendance.objects.filter(attendancedate=target_date)

        #attendance_records = Team.objects.all()

        work = workon.objects.all()

        unique_teamnames = attendance_records.values('Teamname').distinct()

        absent_list = ['Leave', 'CompOff', 'Holiday', 'WeekOff']

        wb = Workbook()
        ws = wb.active
        ws.title="TeamReport"
        shift = wb.create_sheet(title="NewSheet")
        shift.title = "ShiftReport"
        consolidate = wb.create_sheet(title="Consolidate")
        consolidate.title = "ConsolidateReport"
        channelpartner = wb.create_sheet(title="ChannelPartner")
        channelpartner.title="ChannelPartner"
        qc = wb.create_sheet(title="QC")
        qc.title = "QC"


        today = date.today()

        #=======================Header Starts=================================
        #ws.merge_cells('A1:F1')
        ws['A1'] = 'Team Attendance'
        ws['A1'].font = Font(bold=True)

       #=====================Header Ends===============================


       #=======================Row and column headings==============
        ws.cell(row=2, column=1, value='Status')

        col = 2
        for teamname in unique_teamnames:
            ws.cell(row=2, column=col, value=teamname["Teamname"])
            col = col + 1

        row = 3
        for result in work:
            if result.worktype not in absent_list:
                ws.cell(row=row, column=1, value=result.worktype)
                row = row + 1

     #==================Row and column headings ends======================

     #==================Row and column data adding========================

        col = 2
        row = 3
        sumtotsupport=0
        for teamname in unique_teamnames:
            total = 0
            for worktyp in work:

                if worktyp.worktype not in absent_list:

                    filtered_records = DailyAttendance.objects.filter(
                        attendancedate=target_date,
                        Teamname=teamname["Teamname"],
                        worktype=worktyp.worktype
                    )

                    noofrecords = filtered_records.count()
                    total = total + noofrecords

                    if noofrecords==0:
                        ws.cell(row=row, column=col, value="")
                    else:
                        ws.cell(row=row, column=col, value=noofrecords)

                    row = row + 1

            ws.cell(row=row + 1, column=1, value="TotalSupport")
            count_cell = ws.cell(row=row + 1, column=1)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

            ws.cell(row=row + 1, column=col, value=total)
            sumtotsupport= sumtotsupport+total
            count_cell = ws.cell(row=row + 1, column=col)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col+1)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col + 2)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")

            col = col + 1
            row = 3

        row = 3
        ws.cell(row=2, column=col + 1, value="Total")
        count_cell = ws.cell(row=2, column=col + 1)
        count_cell.font = Font(color="000000", bold=True)

        for worktyp in work:

            if worktyp.worktype not in absent_list:

                filtered_records = DailyAttendance.objects.filter(
                    attendancedate=target_date,
                    worktype=worktyp.worktype
                )

                if filtered_records.count()==0:
                    print(worktyp.worktype)
                    ws.delete_rows(row)
                else:
                    ws.cell(row=row, column=col + 1, value=filtered_records.count())
                    count_cell = ws.cell(row=row, column=col + 1)
                    count_cell.font = Font(color="000000", bold=True)
                    row = row + 1


        ws.cell(row=row+1, column=col + 1, value=sumtotsupport)

        #========================Leave/absent========================

        row=row+3
        sumtotabsent = 0
        for result in work:
            if result.worktype in absent_list:
                ws.cell(row=row, column=1, value=result.worktype)
                row = row + 1

        col = 2
        row = row-4
        for teamname in unique_teamnames:
            total = 0
            for worktyp in work:
                if worktyp.worktype in absent_list:
                    filtered_records = DailyAttendance.objects.filter(
                        attendancedate=target_date,
                        Teamname=teamname["Teamname"],
                        worktype=worktyp.worktype
                    )

                    noofrecords = filtered_records.count()
                    total = total + noofrecords

                    if noofrecords==0:
                        ws.cell(row=row, column=col, value="")
                    else:
                        ws.cell(row=row, column=col, value=noofrecords)

                    row = row + 1

            ws.cell(row=row + 1, column=1, value="TotalAbsent")
            count_cell = ws.cell(row=row + 1, column=1)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col+1)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col+2)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
            sumtotabsent = sumtotabsent+total

            col = col + 1
            row = row-4

        for worktyp in work:

            if worktyp.worktype in absent_list:

                filtered_records = DailyAttendance.objects.filter(
                    attendancedate=target_date,
                    worktype=worktyp.worktype
                )

                ws.cell(row=row, column=col + 1, value=filtered_records.count())
                count_cell = ws.cell(row=row, column=col + 1)
                count_cell.font = Font(color="000000", bold=True)
                row = row + 1

        ws.cell(row=row+1, column=col + 1, value=sumtotabsent)

        #===============================Actual========================================

        newrow = row + 4
        newcol = 1
        ws.cell(row=newrow, column=newcol, value="Actual")
        count_cell = ws.cell(row=newrow, column=1)
        count_cell.font = Font(color="000000", bold=True)
        count_cell.fill = PatternFill(start_color="E9F7EF", end_color="E9F7EF", fill_type="solid")
        sumtotalactual = 0

        for teamname in unique_teamnames:

            actualtotal = Team.objects.filter(Teamname=teamname["Teamname"]).count()

            ws.cell(row=newrow, column=newcol + 1).value = actualtotal

            sumtotalactual = sumtotalactual + actualtotal

            #presenttotal = 0
            #absenttotal = 0
            #total=0

            #for worktyp in work:
            #    if worktyp.worktype not in absent_list:
            #        filtered_records = DailyAttendance.objects.filter(
            #            attendancedate=target_date,
            #            Teamname=teamname["Teamname"],
            #            worktype=worktyp.worktype
            #        )
            #        noofrecords = filtered_records.count()
            #        presenttotal = presenttotal + noofrecords

            #for worktyp in work:
            #    if worktyp.worktype in absent_list:
            #        filtered_records = DailyAttendance.objects.filter(
            #            attendancedate=target_date,
            #            Teamname=teamname["Teamname"],
            #            worktype=worktyp.worktype
            #        )
            #        noofrecords = filtered_records.count()
            #        absenttotal = absenttotal + noofrecords

            #total = presenttotal+absenttotal
            #sumtotalactual= sumtotalactual+total

            #ws.cell(row=newrow, column=newcol+1).value = total


            count_cell = ws.cell(row=newrow, column=newcol+1)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="E9F7EF", end_color="E9F7EF", fill_type="solid")

            count_cell = ws.cell(row=newrow, column=newcol + 2)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="E9F7EF", end_color="E9F7EF", fill_type="solid")

            count_cell = ws.cell(row=newrow, column=newcol + 3)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="E9F7EF", end_color="E9F7EF", fill_type="solid")


            fill = PatternFill(start_color="EFEFAF", end_color="EFEFAF", fill_type="solid")
            for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=newcol + 3):
                for cell in row:
                    cell.fill = fill

            for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=newcol + 3):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            newcol = newcol+1

        ws.cell(row=newrow, column=newcol + 2).value = sumtotalactual
        column_number = 1
        column = str(chr(64 + column_number))
        ws.column_dimensions[column].width = 18

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        # Apply borders to rows (e.g., rows 2 to 4)
        for row in range(1, newrow+1):
            for col in range(1, newcol + 3):
                ws.cell(row=row, column=col).border = border_style


        max_column = ws.max_column

        for i in range(2,max_column):
            column = str(chr(64 + i))
            ws.column_dimensions[column].width = 12


        start_column_letter = get_column_letter(1)
        end_column_letter = get_column_letter(max_column)

        ws.merge_cells(f'{start_column_letter}1:{end_column_letter}1')


        self.shiftwisereport(shift,consolidate,channelpartner,qc)

        response = HttpResponse(content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename="Report_{}.xlsx"'.format(today.strftime("%d-%m-%Y"))
        wb.save(response)

        return response

    def shiftwisereport(self,ws,consolidate,channelpartner,qc):

        #absent_list = ['Leave', 'CompOff', 'Holiday']
        absent_list = ['Leave', 'CompOff', 'Holiday', 'WeekOff']

        shift = Shift(attendancedate=self.target_date_str)
        shifttimings = shift.GetShiftTimings()

        #shifttimings = ShiftTimings.objects.all()

        teams_records = Team.objects.all()

        #unique_teamnames = teams_records.values('Teamname').distinct()

        target_date = datetime.strptime(self.target_date_str, "%d/%m/%Y")
        attendance_records = DailyAttendance.objects.filter(attendancedate=target_date)
        unique_teamnames = attendance_records.values('Teamname').distinct()

        ws.cell(row=2, column=1, value='Status')
        count_cell = ws.cell(row=2, column=1)
        count_cell.font = Font(color="000000", bold=True)

        col = 2
        for shift in shifttimings:
            ws.cell(row=2, column=col, value=shift)
            col = col + 1

        row = 3
        for team in unique_teamnames:
            ws.cell(row=row, column=1, value=team['Teamname'])
            row = row + 1


        col = 2
        row = 3
        for shift in shifttimings:
            total = 0
            for team in unique_teamnames:

                filtered_records = DailyAttendance.objects.filter(
                    attendancedate=target_date,
                    Teamname=team['Teamname'],
                    Shift=shift
                )

                count = 0
                for result in filtered_records:
                    if result.worktype not in absent_list and result.worktype not in ('', None) :
                        count = count + 1

                noofrecords = count

                total = total + noofrecords

                if noofrecords !=0:
                    ws.cell(row=row, column=col, value=noofrecords)
                else:
                    ws.cell(row=row, column=col, value="")

                row = row + 1

            ws.cell(row=row + 1, column=1, value="Total")
            count_cell = ws.cell(row=row + 1, column=1)
            count_cell.font = Font(color="000000", bold=True)

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col)
            count_cell.font = Font(color="000000", bold=True)

            col = col + 1
            row = 3

        row = 3
        ws.cell(row=2, column=col, value="Total")
        count_cell = ws.cell(row=2, column=col)
        count_cell.font = Font(color="000000", bold=True)
        totshiftreport=0

        for team in unique_teamnames:

            filtered_records = DailyAttendance.objects.filter(
                attendancedate=target_date,
                Teamname = team["Teamname"]
            )

            count = 0
            for result in filtered_records:
                if result.worktype not in ('', None) and result.worktype not in absent_list:
                    count = count + 1

            ws.cell(row=row, column=col, value=count)
            totshiftreport=totshiftreport+count
            count_cell = ws.cell(row=row, column=col)
            count_cell.font = Font(color="000000", bold=True)
            row = row + 1


        ws.cell(row=row+1, column=col, value=totshiftreport)
        count_cell = ws.cell(row=row+1, column=col)
        count_cell.font = Font(color="000000", bold=True)

        fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        for row in ws.iter_rows(min_row=row+1, max_row=row+1, min_col=1, max_col=col):
            for cell in row:
                cell.fill = fill


        column_number = 1
        column = str(chr(64 + column_number))
        ws.column_dimensions[column].width = 20

        for i in range(2,9):
            column_number = i
            column = str(chr(64 + column_number))
            ws.column_dimensions[column].width = 20


        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        borderlastrow = 0
        for i in range(2,100):
            if str(ws.cell(row=i, column=1).value).strip()=="Total":
                borderlastrow=i
                break

        borderlastcol=0
        for i in range(2, 1000):
            if str(ws.cell(row=2, column=i).value).strip() == "Total":
                borderlastcol = i
                break

        for row in range(2, borderlastrow+1):
            for col in range(1, borderlastcol+1):
                ws.cell(row=row, column=col).border = border_style

        end_column_letter = get_column_letter(col)

        ws.merge_cells(f'A1:{end_column_letter}1')
        ws['A1'] = 'ShiftReport-Present'
        ws['A1'].font = Font(bold=True)


        fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.fill = fill


        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')


       #==============================================ShiftReport-Absent=====================================================

        last_row_number = ws.max_row+5

       #absent_list = ['Leave', 'CompOff', 'Holiday']
        absent_list = ['Leave', 'CompOff', 'Holiday', 'WeekOff']

        #shifttimings = ShiftTimings.objects.all()

        shift = Shift(attendancedate=self.target_date_str)
        shifttimings = shift.GetShiftTimings()

        teams_records = Team.objects.all()

        target_date = datetime.strptime(self.target_date_str, "%d/%m/%Y")
        attendance_records = DailyAttendance.objects.filter(attendancedate=target_date)
        unique_teamnames = attendance_records.values('Teamname').distinct()

        ws.cell(row=last_row_number+1, column=1, value='Status')
        count_cell = ws.cell(row=last_row_number+1, column=1)
        count_cell.font = Font(color="000000", bold=True)

        col = 2
        for shift in shifttimings:
            ws.cell(row=last_row_number+1, column=col, value=shift)
            col = col + 1

        row = last_row_number+2
        for team in unique_teamnames:
            ws.cell(row=row, column=1, value=team['Teamname'])
            row = row + 1

        col = 2
        row = last_row_number+2
        for shift in shifttimings:
            total = 0
            for team in unique_teamnames:
                filtered_records = DailyAttendance.objects.filter(
                    attendancedate=target_date,
                    Teamname=team['Teamname'],
                    Shift=shift
                )

                count = 0
                for result in filtered_records:
                    if result.worktype in absent_list and result.worktype not in ('', None):
                        count = count + 1

                noofrecords = count

                total = total + noofrecords

                if noofrecords != 0:
                    ws.cell(row=row, column=col, value=noofrecords)
                else:
                    ws.cell(row=row, column=col, value="")

                row = row + 1

            ws.cell(row=row + 1, column=1, value="Total")
            count_cell = ws.cell(row=row + 1, column=1)
            count_cell.font = Font(color="000000", bold=True)

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col)
            count_cell.font = Font(color="000000", bold=True)

            col = col + 1
            row = last_row_number+2


        row = last_row_number+2
        ws.cell(row=last_row_number+1, column=col, value="Total")
        count_cell = ws.cell(row=last_row_number+1, column=col)
        count_cell.font = Font(color="000000", bold=True)
        sumtotcount = 0

        for team in unique_teamnames:

            filtered_records = DailyAttendance.objects.filter(
                attendancedate=target_date,
                Teamname=team["Teamname"]
            )

            count = 0
            for result in filtered_records:
                if result.worktype not in ('', None) and result.worktype in absent_list:
                    count = count + 1

            sumtotcount=sumtotcount+count
            ws.cell(row=row, column=col, value=count)
            count_cell = ws.cell(row=row, column=col)
            count_cell.font = Font(color="000000", bold=True)
            row = row + 1

        ws.cell(row=row+1, column=col, value=sumtotcount)
        count_cell = ws.cell(row=row + 1, column=col)
        count_cell.font = Font(color="000000", bold=True)

        fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        for row in ws.iter_rows(min_row=row+1, max_row=row+1, min_col=1, max_col=col):
            for cell in row:
                cell.fill = fill


        borderlastrow = 0
        for i in range(2, 100):
            if str(ws.cell(row=i, column=1).value).strip() == "Total":
                borderlastrow = i

        borderstartrow = 0
        for i in range(2, 100):
            if str(ws.cell(row=i, column=1).value).strip() == "Status":
                borderstartrow = i

        borderlastcol = 0
        for i in range(2, 100):
            if str(ws.cell(row=2, column=i).value).strip() == "Total":
                borderlastcol = i

        for row in range(borderstartrow, borderlastrow + 1):
            for col in range(1, borderlastcol + 1):
                ws.cell(row=row, column=col).border = border_style

        end_column_letter = get_column_letter(col)

        ws.merge_cells(f'A{last_row_number}:{end_column_letter}{last_row_number}')
        ws[f'A{last_row_number}'] = 'ShiftReport-Absent'
        ws[f'A{last_row_number}'].font = Font(bold=True)

        fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        for row in ws.iter_rows(min_row=last_row_number, max_row=last_row_number, min_col=1, max_col=3):
            for cell in row:
                cell.fill = fill

        for row in ws.iter_rows(min_row=last_row_number, max_row=last_row_number, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')


        self.consolidatedreport(consolidate,channelpartner,qc)


    def consolidatedreport(self,ws,channelpartner,qc):

        target_date = datetime.strptime(self.target_date_str, "%d/%m/%Y")

        ws.merge_cells('A1:I1')
        ws['A1'] = 'ConsolidateReport'
        ws['A1'].font = Font(bold=True)

        fill = PatternFill(start_color="EFEFAF", end_color="EFEFAF", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
            for cell in row:
                cell.fill = fill

        fill = PatternFill(start_color="F4F6F6", end_color="F4F6F6", fill_type="solid")
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=9):
            for cell in row:
                cell.fill = fill

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=2, column=1, value='Date')
        ws.cell(row=2, column=2, value='Emp ID')
        ws.cell(row=2, column=3, value='EmpName')
        ws.cell(row=2, column=4, value='Team')
        ws.cell(row=2, column=5, value='SAPID')
        ws.cell(row=2, column=6, value='Shift')
        ws.cell(row=2, column=7, value='TLName')
        ws.cell(row=2, column=8, value='Work')
        ws.cell(row=2, column=9, value='Remarks')

        ws['A2'].font = Font(bold=True)
        ws['B2'].font = Font(bold=True)
        ws['C2'].font = Font(bold=True)
        ws['D2'].font = Font(bold=True)
        ws['E2'].font = Font(bold=True)
        ws['F2'].font = Font(bold=True)
        ws['G2'].font = Font(bold=True)
        ws['H2'].font = Font(bold=True)
        ws['I2'].font = Font(bold=True)


        logins = Login.objects.all()
        row = 3
        for login in logins:
            filtered_records = DailyAttendance.objects.filter(Tlmail=login.username,attendancedate=target_date)
            for record in filtered_records:
                ws.cell(row=row, column=1, value=record.attendancedate.strftime("%d/%m/%Y"))
                ws.cell(row=row, column=2, value=record.Employeeno)
                ws.cell(row=row, column=3, value=record.Employeename)
                ws.cell(row=row, column=4, value=record.Teamname)
                ws.cell(row=row, column=5, value=record.Sapid)
                ws.cell(row=row, column=6, value=record.Shift )
                ws.cell(row=row, column=7, value=record.Tlname)
                ws.cell(row=row, column=8, value=record.worktype)
                ws.cell(row=row, column=9, value=record.remarks)
                row=row+1

        for i in range(1,9):
            column_number = i
            column = str(chr(64 + column_number))
            ws.column_dimensions[column].width = 18

        column_index = 1  # Column A
        alignment = Alignment(horizontal='left')
        for row in ws.iter_rows(min_col=column_index, max_col=column_index):
            for cell in row:
                cell.alignment = alignment

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=7):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')


        last_row_num = ws.max_row

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        for row in range(2, last_row_num+1):
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = border_style

        self.ChannelPartner(channelpartner,qc)

    def ChannelPartner(self,ws,qc):

        #absent_list = ['Leave', 'CompOff', 'Holiday']
        absent_list = ['Leave', 'CompOff', 'Holiday', 'WeekOff']

        shift = Shift(attendancedate=self.target_date_str)
        shifttimings = shift.GetShiftTimings()

        #shifttimings = ShiftTimings.objects.all()

        teams_records = Team.objects.all()

        # unique_teamnames = teams_records.values('Teamname').distinct()

        target_date = datetime.strptime(self.target_date_str, "%d/%m/%Y")
        attendance_records = DailyAttendance.objects.filter(attendancedate=target_date)
        unique_teamnames = attendance_records.values('Teamname').distinct()


        #channelpartnertesting
        channelpartner = [
            "HP Supplies",
            "ESG",
            "ESG - B2B",
            "ESG SNS",
            "Networking",
            "TSG - Networking",
            "Software & Security",
            "TSG - Software & Security",
            "TSG - Server & Storage",
            "MSG - Android",
            "MSG - Apple",
            "MSG",
            "SNS"
        ]

        ws.cell(row=2, column=1, value='Status')
        count_cell = ws.cell(row=2, column=1)
        count_cell.font = Font(color="000000", bold=True)

        col = 2
        for shift in shifttimings:
            ws.cell(row=2, column=col, value=shift)
            col = col + 1

        row = 3
        for team in unique_teamnames:
            if team['Teamname'] in channelpartner:
                ws.cell(row=row, column=1, value=team['Teamname'])
                row = row + 1


        col = 2
        row = 3
        for shift in shifttimings:
            total = 0
            for team in unique_teamnames:

                if team['Teamname'] in channelpartner:

                    filtered_records = DailyAttendance.objects.filter(
                        attendancedate=target_date,
                        Teamname=team['Teamname'],
                        Shift=shift
                    )

                    count = 0

                    for result in filtered_records:

                        if result.worktype not in absent_list and result.worktype not in ('', None):
                            count = count + 1



                    noofrecords = count

                    total = total + noofrecords

                    if noofrecords != 0:
                        ws.cell(row=row, column=col, value=noofrecords)
                    else:
                        ws.cell(row=row, column=col, value="")

                    row = row + 1

            ws.cell(row=row + 1, column=1, value="Total")
            count_cell = ws.cell(row=row + 1, column=1)
            count_cell.font = Font(color="000000", bold=True)

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col)
            count_cell.font = Font(color="000000", bold=True)

            col = col + 1
            row = 3

        row = 3

        ws.cell(row=2, column=col, value="Total")
        count_cell = ws.cell(row=2, column=col)
        count_cell.font = Font(color="000000", bold=True)

        weekday = target_date.weekday()

        if weekday==5:
            ws.cell(row=2, column=col + 1, value="Weekoff")
            count_cell = ws.cell(row=2, column=col + 1)
            count_cell.font = Font(color="000000", bold=True)
        else:
            ws.cell(row=2, column=col + 1, value="Leave")
            count_cell = ws.cell(row=2, column=col + 1)
            count_cell.font = Font(color="000000", bold=True)


        ws.cell(row=2, column=col+2, value="Actual")
        count_cell = ws.cell(row=2, column=col+2)
        count_cell.font = Font(color="000000", bold=True)

        totshiftreport = 0

        for team in unique_teamnames:

            if team['Teamname'] in channelpartner:

                filtered_records = DailyAttendance.objects.filter(
                    attendancedate=target_date,
                    Teamname=team["Teamname"]
                )

                count = 0
                for result in filtered_records:
                    if result.worktype not in ('', None) and result.worktype not in absent_list:
                        count = count + 1

                ws.cell(row=row, column=col, value=count)
                totshiftreport = totshiftreport + count
                count_cell = ws.cell(row=row, column=col)
                count_cell.font = Font(color="000000", bold=True)
                row = row + 1

        ws.cell(row=row + 1, column=col, value=totshiftreport)
        count_cell = ws.cell(row=row + 1, column=col)
        count_cell.font = Font(color="000000", bold=True)

        fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        for row in ws.iter_rows(min_row=row + 1, max_row=row + 1, min_col=1, max_col=col+2):
            for cell in row:
                cell.fill = fill

        column_number = 1
        column = str(chr(64 + column_number))
        ws.column_dimensions[column].width = 20

        for i in range(2, 9):
            column_number = i
            column = str(chr(64 + column_number))
            ws.column_dimensions[column].width = 20

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        borderlastrow = 0
        for i in range(2, 100):
            if str(ws.cell(row=i, column=1).value).strip() == "Total":
                borderlastrow = i
                break

        borderlastcol = 0
        for i in range(2, 1000):
            if str(ws.cell(row=2, column=i).value).strip() == "Total":
                borderlastcol = i
                break

        for row in range(2, borderlastrow + 1):
            for col in range(1, borderlastcol + 3):
                ws.cell(row=row, column=col).border = border_style

        end_column_letter = get_column_letter(col)

        ws.merge_cells(f'A1:{end_column_letter}1')
        ws['A1'] = f'{datetime.today().strftime("%d-%B-%Y")} ChannelPartner-ShiftReport-Present'
        ws['A1'].font = Font(bold=True)

        fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.fill = fill

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ==============================================ShiftReport-Absent=====================================================

        last_row_number = ws.max_row + 5

        #absent_list = ['Leave', 'CompOff', 'Holiday']
        absent_list = ['Leave', 'CompOff', 'Holiday', 'WeekOff']

        # shifttimings = ShiftTimings.objects.all()

        shift = Shift(attendancedate=self.target_date_str)
        shifttimings = shift.GetShiftTimings()

        teams_records = Team.objects.all()

        target_date = datetime.strptime(self.target_date_str, "%d/%m/%Y")
        attendance_records = DailyAttendance.objects.filter(attendancedate=target_date)
        unique_teamnames = attendance_records.values('Teamname').distinct()

        ws.cell(row=last_row_number + 1, column=1, value='Status')
        count_cell = ws.cell(row=last_row_number + 1, column=1)
        count_cell.font = Font(color="000000", bold=True)

        col = 2
        for shift in shifttimings:
            ws.cell(row=last_row_number + 1, column=col, value=shift)
            col = col + 1

        row = last_row_number + 2
        for team in unique_teamnames:

            if team['Teamname'] in channelpartner:
                ws.cell(row=row, column=1, value=team['Teamname'])
                row = row + 1

        col = 2
        row = last_row_number + 2
        for shift in shifttimings:
            total = 0
            for team in unique_teamnames:

                if team['Teamname'] in channelpartner:

                    filtered_records = DailyAttendance.objects.filter(
                        attendancedate=target_date,
                        Teamname=team['Teamname'],
                        Shift=shift
                    )

                    count = 0
                    for result in filtered_records:
                        if result.worktype in absent_list and result.worktype not in ('', None):
                            count = count + 1

                    noofrecords = count

                    total = total + noofrecords

                    if noofrecords != 0:
                        ws.cell(row=row, column=col, value=noofrecords)
                    else:
                        ws.cell(row=row, column=col, value="")

                    row = row + 1

            ws.cell(row=row + 1, column=1, value="Total")
            count_cell = ws.cell(row=row + 1, column=1)
            count_cell.font = Font(color="000000", bold=True)

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col)
            count_cell.font = Font(color="000000", bold=True)

            col = col + 1
            row = last_row_number + 2

        row = last_row_number + 2
        ws.cell(row=last_row_number + 1, column=col, value="Total")
        count_cell = ws.cell(row=last_row_number + 1, column=col)
        count_cell.font = Font(color="000000", bold=True)
        sumtotcount = 0

        for team in unique_teamnames:

            if team['Teamname'] in channelpartner:

                filtered_records = DailyAttendance.objects.filter(
                    attendancedate=target_date,
                    Teamname=team["Teamname"]
                )


                count = 0
                for result in filtered_records:
                    if result.worktype not in ('', None) and result.worktype in absent_list:
                        count = count + 1


                sumtotcount = sumtotcount + count
                ws.cell(row=row, column=col, value=count)
                count_cell = ws.cell(row=row, column=col)
                count_cell.font = Font(color="000000", bold=True)
                row = row + 1

        ws.cell(row=row + 1, column=col, value=sumtotcount)
        count_cell = ws.cell(row=row + 1, column=col)
        count_cell.font = Font(color="000000", bold=True)

        fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        for row in ws.iter_rows(min_row=row + 1, max_row=row + 1, min_col=1, max_col=col):
            for cell in row:
                cell.fill = fill

        borderlastrow = 0
        for i in range(2, 100):
            if str(ws.cell(row=i, column=1).value).strip() == "Total":
                borderlastrow = i

        borderstartrow = 0
        for i in range(2, 100):
            if str(ws.cell(row=i, column=1).value).strip() == "Status":
                borderstartrow = i

        borderlastcol = 0
        for i in range(2, 100):
            if str(ws.cell(row=2, column=i).value).strip() == "Total":
                borderlastcol = i

        for row in range(borderstartrow, borderlastrow + 1):
            for col in range(1, borderlastcol + 1):
                ws.cell(row=row, column=col).border = border_style

        end_column_letter = get_column_letter(col)

        ws.merge_cells(f'A{last_row_number}:{end_column_letter}{last_row_number}')
        ws[f'A{last_row_number}'] = 'ChannelPartner-ShiftReport-Absent'
        ws[f'A{last_row_number}'].font = Font(bold=True)

        fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        for row in ws.iter_rows(min_row=last_row_number, max_row=last_row_number, min_col=1, max_col=3):
            for cell in row:
                cell.fill = fill

        for row in ws.iter_rows(min_row=last_row_number, max_row=last_row_number, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        #======================================Re-Work=====================================================

        
        for col in range(1,50):
            if str(ws.cell(row=2, column=col).value).strip() == "Total":
                present_endcol = col
                absent_endcol = col

                count_cell = ws.cell(row=2, column=present_endcol)
                count_cell.font = Font(color="000000", bold=True)
                count_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

                count_cell = ws.cell(row=2, column=present_endcol+1)
                count_cell.font = Font(color="000000", bold=True)
                count_cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

                count_cell = ws.cell(row=2, column=present_endcol + 2)
                count_cell.font = Font(color="000000", bold=True)
                count_cell.fill = PatternFill(start_color="73C2FB", end_color="73C2FB", fill_type="solid")


        for row in range(1,50):
            if str(ws.cell(row=row, column=1).value).strip() == "Total":
                present_endrow = row
                break

        for row in range(1, 50):
            if str(ws.cell(row=row, column=1).value).strip() == "Status":
                absent_startrow = row


        for row in range(1, 50):
            if str(ws.cell(row=row, column=1).value).strip() == "Total":
                absent_endrow = row



        try:

            sumactualtotal = 0
            lastrow =0
            lastcol = 0
            for i in range(3, present_endrow + 1):
                pres_prod = ws.cell(row=i, column=1).value
                if ws.cell(row=i, column=1).value != None:
                    for j in range(absent_startrow + 1, absent_endrow + 1):
                        abs_prod = ws.cell(row=j, column=1).value
                        if pres_prod == abs_prod:
                            ws.cell(row=i, column=present_endcol + 1).value = ws.cell(row=j, column=absent_endcol).value

                            teamname = ws.cell(row=i, column=1).value
                            actualtotal = Team.objects.filter(Teamname=teamname).count()
                            ws.cell(row=i, column=present_endcol + 2).value = actualtotal
                            sumactualtotal = sumactualtotal + actualtotal
                            lastrow = i
                            lastcol = present_endcol+2


            ws.cell(row=lastrow, column=lastcol).value = sumactualtotal

            for j in range(1, 100):
                for i in range(present_endrow + 1, 150):
                    ws.delete_rows(i)

            for i in range(2, 100):
                print(ws.cell(row=2, column=i).value)
                if ws.cell(row=2, column=i).value == "Total":
                    ws.cell(row=2, column=i).value = "Today Team"
                if ws.cell(row=2, column=i).value == "Actual":
                    ws.cell(row=2, column=i).value = "Actual Team"

        except Exception as e:
            print(e)
        


        self.QC(qc)

    def QC(self,ws):

        # absent_list = ['Leave', 'CompOff', 'Holiday']
        absent_list = ['Leave', 'CompOff', 'Holiday', 'WeekOff']

        shift = Shift(attendancedate=self.target_date_str)
        shifttimings = shift.GetShiftTimings()

        # shifttimings = ShiftTimings.objects.all()

        teams_records = Team.objects.all()

        # unique_teamnames = teams_records.values('Teamname').distinct()

        target_date = datetime.strptime(self.target_date_str, "%d/%m/%Y")
        attendance_records = DailyAttendance.objects.filter(attendancedate=target_date)
        unique_teamnames = attendance_records.values('Teamname').distinct()

        channelpartner = [
            "QC Reversal - OBD",
            "QC Report/Reversal",
            "QC - Reversal",
            "QC PR",
            "QC - PR",
            "QC PI",
            "QC - PI",
            "QC OBD",
            "QC - OBD",
            "QC  Report",
            "QC - Report",
            "QC",
            "QC - QC"
        ]

        team_records = Team.objects.all()
        unique_teamnames = team_records.values('Teamname').distinct()

        for team in unique_teamnames:
            if team['Teamname'] in channelpartner:
                qctotal = Team.objects.filter(Teamname=team["Teamname"]).count()

        ws.cell(row=2, column=1, value='Status')
        count_cell = ws.cell(row=2, column=1)
        count_cell.font = Font(color="000000", bold=True)

        col = 2
        for shift in shifttimings:
            ws.cell(row=2, column=col, value=shift)
            col = col + 1

        row = 3
        for team in unique_teamnames:
            if team['Teamname'] in channelpartner:
                ws.cell(row=row, column=1, value=team['Teamname'])
                row = row + 1

        col = 2
        row = 3
        for shift in shifttimings:
            total = 0
            for team in unique_teamnames:

                if team['Teamname'] in channelpartner:

                    filtered_records = DailyAttendance.objects.filter(
                        attendancedate=target_date,
                        Teamname=team['Teamname'],
                        Shift=shift
                    )

                    count = 0
                    for result in filtered_records:
                        if result.worktype not in absent_list and result.worktype not in ('', None):
                            count = count + 1

                    noofrecords = count

                    total = total + noofrecords

                    if noofrecords != 0:
                        ws.cell(row=row, column=col, value=noofrecords)
                    else:
                        ws.cell(row=row, column=col, value="")

                    row = row + 1

            ws.cell(row=row + 1, column=1, value="Total")
            count_cell = ws.cell(row=row + 1, column=1)
            count_cell.font = Font(color="000000", bold=True)

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col)
            count_cell.font = Font(color="000000", bold=True)

            col = col + 1
            row = 3

        row = 3
        ws.cell(row=2, column=col, value="Total")
        count_cell = ws.cell(row=2, column=col)
        count_cell.font = Font(color="000000", bold=True)
        totshiftreport = 0

        for team in unique_teamnames:

            if team['Teamname'] in channelpartner:

                print(team["Teamname"])

                filtered_records = DailyAttendance.objects.filter(
                    attendancedate=target_date,
                    Teamname=team["Teamname"]
                )

                count = 0
                for result in filtered_records:
                    if result.worktype not in ('', None) and result.worktype not in absent_list:
                        count = count + 1

                ws.cell(row=row, column=col, value=count)
                totshiftreport = totshiftreport + count
                count_cell = ws.cell(row=row, column=col)
                count_cell.font = Font(color="000000", bold=True)
                row = row + 1

        ws.cell(row=row + 1, column=col, value=totshiftreport)
        count_cell = ws.cell(row=row + 1, column=col)
        count_cell.font = Font(color="000000", bold=True)

        fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        for row in ws.iter_rows(min_row=row + 1, max_row=row + 1, min_col=1, max_col=col):
            for cell in row:
                cell.fill = fill

        column_number = 1
        column = str(chr(64 + column_number))
        ws.column_dimensions[column].width = 20

        for i in range(2, 9):
            column_number = i
            column = str(chr(64 + column_number))
            ws.column_dimensions[column].width = 20

        border_style = Border(left=Side(style='thin'),
                              right=Side(style='thin'),
                              top=Side(style='thin'),
                              bottom=Side(style='thin'))

        borderlastrow = 0
        for i in range(2, 100):
            if str(ws.cell(row=i, column=1).value).strip() == "Total":
                borderlastrow = i
                break

        borderlastcol = 0
        for i in range(2, 1000):
            if str(ws.cell(row=2, column=i).value).strip() == "Total":
                borderlastcol = i
                break

        for row in range(2, borderlastrow + 1):
            for col in range(1, borderlastcol + 1):
                ws.cell(row=row, column=col).border = border_style

        end_column_letter = get_column_letter(col)

        ws.merge_cells(f'A1:{end_column_letter}1')
        ws['A1'] = 'QC-ShiftReport-Present'
        ws['A1'].font = Font(bold=True)

        fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.fill = fill

        for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # ==============================================ShiftReport-Absent=====================================================

        last_row_number = ws.max_row + 5

        # absent_list = ['Leave', 'CompOff', 'Holiday']
        absent_list = ['Leave', 'CompOff', 'Holiday', 'WeekOff']

        # shifttimings = ShiftTimings.objects.all()

        shift = Shift(attendancedate=self.target_date_str)
        shifttimings = shift.GetShiftTimings()

        teams_records = Team.objects.all()

        target_date = datetime.strptime(self.target_date_str, "%d/%m/%Y")
        attendance_records = DailyAttendance.objects.filter(attendancedate=target_date)
        unique_teamnames = attendance_records.values('Teamname').distinct()

        ws.cell(row=last_row_number + 1, column=1, value='Status')
        count_cell = ws.cell(row=last_row_number + 1, column=1)
        count_cell.font = Font(color="000000", bold=True)

        col = 2
        for shift in shifttimings:
            ws.cell(row=last_row_number + 1, column=col, value=shift)
            col = col + 1

        row = last_row_number + 2
        for team in unique_teamnames:

            if team['Teamname'] in channelpartner:
                ws.cell(row=row, column=1, value=team['Teamname'])
                row = row + 1

        col = 2
        row = last_row_number + 2
        for shift in shifttimings:
            total = 0
            for team in unique_teamnames:

                if team['Teamname'] in channelpartner:

                    filtered_records = DailyAttendance.objects.filter(
                        attendancedate=target_date,
                        Teamname=team['Teamname'],
                        Shift=shift
                    )

                    count = 0
                    for result in filtered_records:
                        if result.worktype in absent_list and result.worktype not in ('', None):
                            count = count + 1

                    noofrecords = count

                    total = total + noofrecords

                    if noofrecords != 0:
                        ws.cell(row=row, column=col, value=noofrecords)
                    else:
                        ws.cell(row=row, column=col, value="")

                    row = row + 1

            ws.cell(row=row + 1, column=1, value="Total")
            count_cell = ws.cell(row=row + 1, column=1)
            count_cell.font = Font(color="000000", bold=True)

            ws.cell(row=row + 1, column=col, value=total)
            count_cell = ws.cell(row=row + 1, column=col)
            count_cell.font = Font(color="000000", bold=True)

            col = col + 1
            row = last_row_number + 2

        row = last_row_number + 2
        ws.cell(row=last_row_number + 1, column=col, value="Total")
        count_cell = ws.cell(row=last_row_number + 1, column=col)
        count_cell.font = Font(color="000000", bold=True)
        sumtotcount = 0

        for team in unique_teamnames:

            if team['Teamname'] in channelpartner:

                filtered_records = DailyAttendance.objects.filter(
                    attendancedate=target_date,
                    Teamname=team["Teamname"]
                )

                count = 0
                for result in filtered_records:
                    if result.worktype not in ('', None) and result.worktype in absent_list:
                        count = count + 1

                sumtotcount = sumtotcount + count
                ws.cell(row=row, column=col, value=count)
                count_cell = ws.cell(row=row, column=col)
                count_cell.font = Font(color="000000", bold=True)
                row = row + 1

        ws.cell(row=row + 1, column=col, value=sumtotcount)
        count_cell = ws.cell(row=row + 1, column=col)
        count_cell.font = Font(color="000000", bold=True)

        fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        for row in ws.iter_rows(min_row=row + 1, max_row=row + 1, min_col=1, max_col=col):
            for cell in row:
                cell.fill = fill

        borderlastrow = 0
        for i in range(2, 100):
            if str(ws.cell(row=i, column=1).value).strip() == "Total":
                borderlastrow = i

        borderstartrow = 0
        for i in range(2, 100):
            if str(ws.cell(row=i, column=1).value).strip() == "Status":
                borderstartrow = i

        borderlastcol = 0
        for i in range(2, 100):
            if str(ws.cell(row=2, column=i).value).strip() == "Total":
                borderlastcol = i

        for row in range(borderstartrow, borderlastrow + 1):
            for col in range(1, borderlastcol + 1):
                ws.cell(row=row, column=col).border = border_style

        end_column_letter = get_column_letter(col)

        ws.merge_cells(f'A{last_row_number}:{end_column_letter}{last_row_number}')
        ws[f'A{last_row_number}'] = 'QC-ShiftReport-Absent'
        ws[f'A{last_row_number}'].font = Font(bold=True)

        fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")
        for row in ws.iter_rows(min_row=last_row_number, max_row=last_row_number, min_col=1, max_col=3):
            for cell in row:
                cell.fill = fill

        for row in ws.iter_rows(min_row=last_row_number, max_row=last_row_number, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')

        # =====================================qc report consolidation===============================

        try:
            lastrow = 0
            for i in range(1, 1000):
                qcval = str(ws.cell(row=i, column=1).value).strip()
                if qcval == "Total":
                    last_row_number = i + 3

            lastcol = 0
            for i in range(1, 1000):
                qcval = str(ws.cell(row=2, column=i).value).strip()
                if qcval == "Total":
                    lastcol = i

            end_column_letter = get_column_letter(lastcol + 2)

            ws.merge_cells(f'A{last_row_number}:{end_column_letter}{last_row_number}')

            ws[f'A{last_row_number}'] = f'{datetime.today().strftime("%d-%B-%Y")} QC-ShiftReport-Present'
            ws[f'A{last_row_number}'].font = Font(bold=True)

            fill = PatternFill(start_color="EAFAF1", end_color="EAFAF1", fill_type="solid")
            for row in ws.iter_rows(min_row=last_row_number, max_row=last_row_number, min_col=1, max_col=3):
                for cell in row:
                    cell.fill = fill

            for row in ws.iter_rows(min_row=last_row_number, max_row=last_row_number, min_col=1, max_col=3):
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            ws.cell(row=last_row_number + 1, column=1).value = "Status"
            ws.cell(row=last_row_number + 1, column=1).font = Font(bold=True)
            ws.cell(row=last_row_number + 2, column=1).value = "QC"

            for i in range(2, lastcol):
                ws.cell(row=last_row_number + 1, column=i).value = ws.cell(row=2, column=i).value

            ws.cell(row=last_row_number + 1, column=i + 1).value = "Today Team"
            count_cell = ws.cell(row=last_row_number + 1, column=i + 1)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

            weekday = target_date.weekday()

            if weekday==5:
                ws.cell(row=last_row_number + 1, column=i + 2).value = "Weekoff"
                count_cell = ws.cell(row=last_row_number + 1, column=i + 2)
                count_cell.font = Font(color="000000", bold=True)
                count_cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")
            else:
                ws.cell(row=last_row_number + 1, column=i + 2).value = "Leave"
                count_cell = ws.cell(row=last_row_number + 1, column=i + 2)
                count_cell.font = Font(color="000000", bold=True)
                count_cell.fill = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")


            ws.cell(row=last_row_number + 1, column=i + 3).value = "Actual Team"
            count_cell = ws.cell(row=last_row_number + 1, column=i + 3)
            count_cell.font = Font(color="000000", bold=True)
            count_cell.fill = PatternFill(start_color="73C2FB", end_color="73C2FB", fill_type="solid")

            border_style = Border(left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))

            for row in range(last_row_number + 1, last_row_number + 3):
                for col in range(1, i + 4):
                    ws.cell(row=row, column=col).border = border_style

            # Writeout of calculated value in qc-shift-report-summary table

            qc_present_total_row = 0
            for i in range(1, 1000):
                if ws.cell(row=i, column=1).value == "Total":
                    qc_present_total_row = i
                    break

            qc_present_last_col = 0
            for i in range(1, 1000):
                qcval = str(ws.cell(row=2, column=i).value).strip()
                if qcval == "Total":
                    qc_present_last_col = i

            qc_summary_row = 0
            for i in range(1, 1000):
                qcval = str(ws.cell(row=i, column=1).value).strip()
                if qcval == "QC":
                    qc_summary_row = i

            for i in range(2, qc_present_last_col + 1):
                qcval = ws.cell(row=qc_present_total_row, column=i).value
                ws.cell(row=qc_summary_row, column=i).value = qcval

            qc_absent_total_row = 0
            for i in range(1, 1000):
                if ws.cell(row=i, column=1).value == "Total":
                    qc_absent_total_row = i

            absent_total_val = ws.cell(row=qc_absent_total_row, column=qc_present_last_col).value

            ws.cell(row=qc_summary_row, column=qc_present_last_col + 1).value = absent_total_val
            # ws.cell(row=qc_summary_row, column=qc_present_last_col + 2).value = ws.cell(row=qc_summary_row, column=qc_present_last_col).value+ absent_total_val

            ws.cell(row=qc_summary_row, column=qc_present_last_col + 2).value = qctotal

            # Deleting columns which is having 0 as total

            qclastrow = ""
            for num in range(1000, 0, -1):
                if str(ws.cell(row=num, column=1).value).strip() == "QC":
                    qclastrow = num
                    break

            qclastcol = ""
            for i in range(1, 100):
                if str(ws.cell(row=qclastrow - 1, column=i).value).strip() == "Actual Team":
                    qclastcol = i
                    break

            for loop in range(1, 4):
                for i in range(1, qclastcol - 2):
                    if ws.cell(row=qclastrow, column=i).value == 0 and str(
                            ws.cell(row=qclastrow - 1, column=i).value) not in ["Today Team", "Leave","Weekoff",
                                                                                "Actual Team"]:
                        pass
                        ws.delete_cols(i)

            # Unmerge and Merge cells

            lastcol = ""
            for i in range(1, 100):
                if str(ws.cell(row=2, column=i).value) == "Total":
                    lastcol = i
                    break

            old_column_name_summary = get_column_letter(qclastcol)
            new_column_name_summary = get_column_letter(lastcol + 2)
            old_column_name_general = get_column_letter(qclastcol - 2)
            new_column_name_general = get_column_letter(lastcol)

            print(old_column_name_summary)
            print(old_column_name_general)
            print(new_column_name_summary)
            print(new_column_name_general)

            range_general = f'A1:{old_column_name_general}1'
            ws.unmerge_cells(range_general)

            range_general = f'A1:{new_column_name_general}1'
            ws.merge_cells(range_general)

            rowalign = ""
            for i in range(1, 100):
                if str(ws.cell(row=i, column=1).value) == "QC-ShiftReport-Absent":
                    rowalign = i
                    break

            range_general = f'A{rowalign}:{old_column_name_general}{rowalign}'
            ws.unmerge_cells(range_general)

            range_general = f'A{rowalign}:{new_column_name_general}{rowalign}'
            ws.merge_cells(range_general)

            rowalign = ""
            for i in range(100, 0, -1):
                if "QC-ShiftReport-Present" in str(ws.cell(row=i, column=1).value):
                    rowalign = i
                    break

            range_general = f'A{rowalign}:{old_column_name_summary}{rowalign}'
            ws.unmerge_cells(range_general)

            range_general = f'A{rowalign}:{new_column_name_summary}{rowalign}'
            ws.merge_cells(range_general)


        except Exception as e:
            print(e)



































